"""Extract the pre-compilation Laurentia poles into a CSV in the Nordic layout.

The compilation is assembled from two files:

- ``data/nordic_summaries/nordic_summaries_combined.csv`` -- poles at or below
  ``PUBLISH_AGE_MAX``, recomputed from site-level data by this project and
  documented in a notebook apiece.
- ``data/older_Laurentia_poles.csv`` -- this file's output. Poles older than
  that cutoff, which have *not* been recreated at the site level here and are
  carried over from the Nordic workshop compilation unchanged.

Both carry the same 71 columns in the same order, so concatenating them gives
the full record without any reshaping.

The rows are taken from the workshop workbook rather than from this project's
own copy of the earlier compilation (``data/Laurentia_poles.csv``, a 68-column
layout) because the workbook is already in the target layout: no column mapping
step means no column mapping to get wrong. This script is therefore run when the
workshop compilation is updated, not on every build; its output is committed.

    python scripts/extract_older_poles.py [path/to/workshop_workbook.xlsx]
"""

import csv
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SUMMARY_CSV = os.path.join(ROOT, "data", "nordic_summaries",
                           "nordic_summaries_combined.csv")
BUILD_POLE_MAP = os.path.join(HERE, "build_pole_map.py")
OUT_CSV = os.path.join(ROOT, "data", "older_Laurentia_poles.csv")

DEFAULT_WORKBOOK = os.path.expanduser(
    "~/Downloads/Iloranta-global-2026-06-24.xlsx")

# Sampling localities absent from the workshop workbook, ROCKNAME -> (lat, lon
# in degrees east). A pole without one is still tabulated but cannot be placed
# on the interactive map, so the coordinate is supplied here rather than left
# blank. Each entry records where the value comes from.
#
#   Elbow Creek dikes -- mean of the thirteen component-A dikes of Ding et al.
#   (2024, Nature Communications 15:10814, doi:10.1038/s41467-024-55117-w)
#   Supplementary Table S1, the same thirteen that make their pole. They span
#   45.378-45.474 deg N and 109.878-110.139 deg W in the Stillwater Complex of
#   the Beartooth Mountains, Montana, giving a mean of 45.42 deg N, 109.96 deg W
#   (250.04 deg E). As a check, the VGP of their tilt-corrected mean direction
#   (D = 145.8, I = 59.6) at this site is 1.2 N / 275.4 E against their published
#   2.0 N / 275.3 E -- the 0.8 deg offset expected between the VGP of the mean
#   direction and the mean of the thirteen site VGPs that the pole actually is.
SITE_COORD_OVERRIDES = {
    "Elbow Creek dikes": (45.42, 250.04),
}

# Corrections to values carried in the workshop workbook, ROCKNAME -> {column
# header: (workbook value, corrected value)}. Every column sharing the header is
# corrected, which is what the repeated ``PLONf`` blocks need. A correction is
# applied only where the workbook still holds the stale value; once it is fixed
# upstream the entry raises rather than silently masking a later revision, so
# each round of workbook corrections shortens this list. Each entry records the
# evidence for the correction.
#
#   Indin dykes -- Buchan et al. (2016, Precambrian Research 275:151-175) give
#   the pole as 36 N, 76 W in their text (p. 158) and their Table 2 mean, and as
#   "36, 284" in degrees east in their APW path table (their Table 5). The
#   workbook carries 76 in a column that is degrees east, placing the pole in
#   central Asia rather than off the US east coast. Reported by D.A.D. Evans,
#   2026-08-25, along with a tightened age bracket (2124-2128 rather than
#   2123-2129, the U-Pb baddeleyite upper intercept quoted as 2126+/-3 or
#   2126+/-2), the spelling of Bleeker, and the lithology. 'age diff' is not set
#   here because the script recomputes it from the bracket.
#
#   Ghost dykes -- the same west-for-east slip in the inclination-corrected
#   columns only. Buchan et al. (2016) give the Ghost pole as 2 N, 106 W
#   (p. 165), which is 254 E; the workbook's PLONG is already 254 and only the
#   two PLONf blocks carry 106. These columns are not read by the analysis --
#   Ghost dykes are igneous with f = 1, so PLONf must equal PLONG -- but the
#   value is wrong where it stands.
#
#   MEAN Pearson A/Peninsular sill/Kilohigok basin sill -- the f blocks carry
#   -13 N, 249 E against a main pole of -22 N, 269 E. This is not a west-for-east
#   slip but a stale pair of values: the unit is igneous with f = 1, so the f
#   blocks must repeat the main pole. The main pole is the one to keep -- it is
#   the VGP of the tabulated direction (162.4/3.5 at 65.0 N, 250.0 E) to within
#   0.05 deg, whereas the f-block values sit 21.1 deg away -- and it is what
#   Evans et al. (2021) carried before the f columns were added.
#
#   Defeat Suite granitoids -- the pole (-1 N, 64 E) resembles the
#   Mesoproterozoic Laurentian APW path, which would normally fail R7 ("no
#   resemblance to paleopoles of younger age"). A positive inverse baked contact
#   test against the ca. 2193 Ma Dogrib dykes (Mitchell et al., 2014, American
#   Journal of Science 314:878-894) constrains the Defeat remanence to predate
#   the Dogrib intrusion, so the resemblance is coincidental and R7 is met.
#   R7 0 -> 1 carries Rsum 2 -> 3. Grade stays B: every pole at Rsum = 3 in this
#   compilation is graded B. Reported by D.A.D. Evans, 2026-08-25.
WORKBOOK_CORRECTIONS = {
    "Indin dykes": {
        "PLONG": (76, 284),
        "PLONf": (76, 284),
        "lomagage": (2123, 2124),
        "himagage": (2129, 2128),
        "REF/method": (
            "2126\u00b13, U-Pb baddeleyite upper intercept ages from Buchan et "
            "al., 2016; although they cite abstracts by Blleker et al., 2008 to "
            "have a much younger date of 2108 Ma, that date is not published "
            "and we do not consider it here. ",
            "2126\u00b13 or 2126\u00b12, U-Pb baddeleyite upper intercept ages "
            "from Buchan et al., 2016; although they cite abstracts by Bleeker "
            "et al., 2008 to have a much younger date of 2108 Ma, that date is "
            "not published and we do not consider it here."),
        "Lithology": (None, "dykes (mafic)"),
    },
    "Ghost dykes": {
        "PLONf": (106, 254),
    },
    "MEAN Pearson A/Peninsular sill/Kilohigok basin sill": {
        "PLATf": (-13, -22),
        "PLONf": (249, 269),
    },
    "Defeat Suite granitoids": {
        "R7": (0, 1),
        "Rsum": (2, 3),
    },
}

# Poles whose workbook entry has been superseded by a later publication that
# reworks the same data, ROCKNAME -> {column header: (superseded value, new
# value)}. Shape and guard behave as in WORKBOOK_CORRECTIONS, and the two are
# applied in order, so a superseded value here is the value *after* any
# correction above. This is kept separate from WORKBOOK_CORRECTIONS because the
# entries mean different things: a correction repairs a transcription error in a
# value the workbook meant to carry, while a supersession swaps in a different,
# later result. Each entry records what supersedes what and why.
#
#   Indin dykes -- Liu et al. (2024, JGR Solid Earth 129:e2024JB029046,
#   doi:10.1029/2024JB029046) add 9 sites in the central and eastern Slave
#   craton to the 18 dykes of Buchan et al. (2016), widening the sampled swath
#   across the craton, and reclassify the combined 28 sites into individual
#   dykes so that dykes sampled at more than one site are not counted more than
#   once (their Table 3). Their consolidated dyke-level mean of 20 dykes is
#   adopted here: D = 116.7, I = 71.7, k = 45.8, a95 = 4.9 at the common
#   reference locality 62.5 N, 114.5 W, giving 40.3 N, 284.7 E, A95 = 8.0.
#
#   The 20-dyke mean is preferred over the 16 SED-only mean of the same table
#   (39.9 N, 284.4 E, A95 = 9.4). Liu et al. prefer the SED-only set for their
#   paleosecular-variation calculation, to keep a possible age difference
#   between the polarity groups from inflating the VGP scatter. That is a
#   concern specific to a scatter calculation, and the case for carrying it into
#   the compiled pole does not hold up:
#
#   - The precisely dated dyke carries an *upward* remanence. Buchan et al.
#     (2016) date site I05 at 2126 +/- 3 Ma on three baddeleyite fractions, and
#     2126 +/- 2 Ma combined with a collinear fraction from site I00 on the same
#     dyke (their Section 6.1); they then state that "the dated site (I05) has an
#     upward Indin remanence direction" (their Section 6.2). I05 is part of NWU2
#     in Liu et al.'s Table 3. Dropping the NWU dykes would therefore discard the
#     polarity group that the tabulated age actually dates.
#   - The suggestion that the swarm spans as much as 20 Myr rests on a ca. 2108
#     Ma U-Pb baddeleyite age reported only in abstracts (Davis & Bleeker, 2007;
#     Bleeker et al., 2008a). This compilation does not use that date, as the
#     REF/method note for this pole records, so it cannot be the basis for
#     partitioning the pole either. That age is in any case from another
#     northwest-trending dyke, not from a southeast-and-down one.
#
#   Retaining both polarities also keeps the positive reversal test attached to
#   the tabulated mean and rests the pole on all 20 independent dykes. The two
#   means lie 0.5 deg apart, so nothing in the APW path turns on the choice, and
#   tabulating both would be redundant rather than informative.
#
#   Geochronology is unchanged: Liu et al. add no new dates, and the age
#   constraint remains that of Buchan et al. (2016). The Q and R criteria and
#   the Grade A rating are likewise unchanged -- the added dykes strengthen the
#   sampling behind R2 without moving any criterion across a threshold. The
#   GPMDB result number is kept, as that result's data are included in the new
#   mean, and the Comment column records the supersession.
SUPERSEDING_RESULTS = {
    "Indin dykes": {
        "SLONG": (245.6, 245.5),
        "B": (18, 20),
        "DEC": (300, 116.7),
        "INC": (-70, 71.7),
        "INCf": (-70, 71.7),
        "KD": (62, 45.8),
        "ED95": (4, 4.9),
        "PLAT": (36, 40.3),
        "PLONG": (284, 284.7),
        "PLATf": (36, 40.3),
        "PLONf": (284, 284.7),
        "DP": (7, 8.0),
        "DM": (7, 8.0),
        "A95": (7, 8.0),
        "DPf": (7, 8.0),
        "DMf": (7, 8.0),
        "A95f": (7, 8.0),
        "%REV": ("33or67", "20or80"),
        "POLE AUTHORS": (
            "Buchan, Kenneth L. and Mitchell, Ross N. and Bleeker, Wouter and "
            "Hamilton, Michael A. and LeCheminant, Anthony N.",
            "Liu, Yu-Shu and Mitchell, Ross N. and Bleeker, Wouter and Peng, "
            "Peng and Salminen, Johanna and Evans, David A. D."),
        "YEAR": (2016, 2024),
        "JOURNAL": ("Precambrian Research",
                    "Journal of Geophysical Research: Solid Earth"),
        "VOLUME": (275, 129),
        "VPAGES": ("151-175", "e2024JB029046"),
        "TITLE": (None,
                  "Conformably Variable Geocentric Axial Dipole at ca. 2.1 Ga: "
                  "Paleomagnetic Dispersion of the Indin Dyke Swarm, Slave "
                  "Craton"),
        "Comment": (None,
                    "Dyke-level mean of 20 dykes (4 northwest-and-up, 16 "
                    "southeast-and-down) from Table 3 of Liu et al. (2024), "
                    "which consolidates dykes sampled at more than one site. "
                    "Supersedes and includes GPMDB result 9484, the 18-dyke "
                    "mean of Buchan et al. (2016). Both polarities are retained, "
                    "so the positive reversal test applies to this mean. The "
                    "SED-only mean of 16 of these dykes, which Liu et al. prefer "
                    "for their paleosecular-variation calculation, is 39.9 N, "
                    "284.4 E, A95 = 9.4 -- indistinguishable from this pole. The "
                    "precisely dated dyke (site I05, 2126 +/- 2 Ma) carries an "
                    "upward remanence, so the dated polarity group is one that "
                    "an SED-only mean would exclude."),
    },
}

# 1-based column indices in the fixed layout.
COL_TERRANE = 1
COL_ROCKNAME = 2
COL_AGE_DIFF = 59
COL_NOMINAL_AGE = 58
COL_LOMAGAGE = 60
COL_HIMAGAGE = 61
COL_SLAT = 8
COL_SLONG = 9


def publish_age_max():
    """The age at which the site-level compilation hands over to older poles.

    Read from ``build_pole_map.py`` rather than duplicated, so the two cannot
    drift apart. That module is not imported because it pulls in the plotting
    stack for a single constant.

    Returns:
        float: The cutoff in Ma.
    """
    with open(BUILD_POLE_MAP, encoding="utf-8") as fh:
        m = re.search(r"^PUBLISH_AGE_MAX\s*=\s*([0-9.]+)", fh.read(), re.M)
    if not m:
        raise SystemExit("PUBLISH_AGE_MAX not found in build_pole_map.py")
    return float(m.group(1))


def cell(value):
    """Render a spreadsheet value for CSV, leaving blanks genuinely empty."""
    return "" if value is None else str(value)


def same_value(a, b):
    """Whether two spreadsheet values are the same number, or the same text.

    Args:
        a: A value read from the workbook, or a value to compare it against.
        b: The other value.

    Returns:
        bool: True if both parse as equal numbers, or are equal as stripped
        strings.
    """
    try:
        return float(a) == float(b)
    except (TypeError, ValueError):
        return str(a).strip() == str(b).strip()


def apply_edits(table, table_name, name, row, header):
    """Applies one edit table's entry for a single pole, in place.

    Shared by WORKBOOK_CORRECTIONS and SUPERSEDING_RESULTS, which have the same
    ``{ROCKNAME: {column: (old value, new value)}}`` shape. Every column sharing
    a header is edited, which is what the repeated ``INCf``/``PLATf``/``PLONf``
    blocks need. An old value of ``None`` skips the guard and replaces whatever
    the workbook holds, for free-text columns not worth pinning exactly.

    Args:
        table (dict): The edit table to apply.
        table_name (str): Its name, used in error messages.
        name (str): The pole's ROCKNAME.
        row (list): The workbook row, modified in place.
        header (list): The column headers, used to locate columns by name.

    Returns:
        list: Human-readable descriptions of the edits applied, empty if the
        pole has no entry in this table.

    Raises:
        SystemExit: If a named column is not in the layout, or the workbook no
            longer holds the value the entry expects -- meaning it has been
            revised upstream, and the entry should be dropped or updated rather
            than left to overwrite the new value.
    """
    edits = table.get(name)
    if not edits:
        return []
    applied = []
    for column, (old, new) in edits.items():
        indices = [i for i, h in enumerate(header) if str(h) == column]
        if not indices:
            raise SystemExit(
                f"{name!r} has a {table_name} entry for column {column!r}, "
                "which is not in the workbook layout.")
        for i in indices:
            if old is not None and not same_value(row[i], old):
                raise SystemExit(
                    f"{name!r} now carries {row[i]!r} in {column!r} rather than "
                    f"the {old!r} its {table_name} entry expects. If it has "
                    "been revised upstream, drop the entry; otherwise update it "
                    "to match.")
            row[i] = new
        shown = f"{old} -> {new}" if old is not None else f"set to {new}"
        applied.append(f"{name}: {column} "
                       + (shown[:80] + "..." if len(shown) > 80 else shown))
    return applied


def main():
    workbook = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not (os.path.exists(workbook) and os.path.getsize(workbook) > 0):
        raise SystemExit(
            f"Workshop workbook not readable at {workbook}. Pass a path as the "
            "first argument, and if it is a cloud placeholder make it available "
            "offline first.")

    ws = openpyxl.load_workbook(workbook).active
    ncol = ws.max_column
    header = [ws.cell(row=1, column=c).value for c in range(1, ncol + 1)]

    with open(SUMMARY_CSV, newline="", encoding="utf-8-sig") as fh:
        our_header = next(csv.reader(fh))
        recreated = {r[COL_ROCKNAME - 1].strip()
                     for r in csv.reader(fh) if len(r) > 1}
    if [str(h) for h in header] != [str(h) for h in our_header]:
        raise SystemExit(
            "The workbook layout does not match the site-level summaries. The "
            "two halves of the compilation must share a layout to concatenate.")

    cutoff = publish_age_max()
    kept, superseded, filled = [], [], []
    corrected, replaced = [], []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ncol + 1)]
        if not str(row[COL_TERRANE - 1]).startswith("Laurentia"):
            continue
        try:
            age = float(row[COL_NOMINAL_AGE - 1])
        except (TypeError, ValueError):
            continue
        if age <= cutoff:
            continue
        name = str(row[COL_ROCKNAME - 1]).strip()
        if name in recreated:
            superseded.append(name)
            continue
        override = SITE_COORD_OVERRIDES.get(name)
        if override is not None:
            if row[COL_SLAT - 1] not in (None, "") \
                    and row[COL_SLONG - 1] not in (None, ""):
                raise SystemExit(
                    f"{name!r} now carries a sampling locality in the workbook; "
                    "drop its SITE_COORD_OVERRIDES entry rather than overwriting "
                    "the workbook value.")
            row[COL_SLAT - 1], row[COL_SLONG - 1] = override
            filled.append(name)
        corrected.extend(apply_edits(WORKBOOK_CORRECTIONS,
                                     "WORKBOOK_CORRECTIONS",
                                     name, row, header))
        replaced.extend(apply_edits(SUPERSEDING_RESULTS, "SUPERSEDING_RESULTS",
                                    name, row, header))
        # 'age diff' is a formula in the workbook, pointing at its row there.
        # Store the value instead, so the CSV stands on its own; the workbook
        # export rewrites it as a formula for the row it lands on.
        try:
            row[COL_AGE_DIFF - 1] = (int(round(float(row[COL_HIMAGAGE - 1])))
                                     - int(round(float(row[COL_LOMAGAGE - 1]))))
        except (TypeError, ValueError):
            row[COL_AGE_DIFF - 1] = None
        kept.append((age, row))

    kept.sort(key=lambda pair: pair[0])
    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(header)
        writer.writerows([cell(v) for v in row] for _, row in kept)

    print(f"Source: {os.path.basename(workbook)}")
    if superseded:
        print("  dropped as superseded by a site-level recreation: "
              + ", ".join(sorted(superseded)))
    if filled:
        print("  sampling locality supplied from the literature: "
              + ", ".join(sorted(filled)))
    if corrected:
        print("  corrected against the source publication:")
        for note in sorted(corrected):
            print(f"    {note}")
    if replaced:
        print("  superseded by a later publication:")
        for note in sorted(replaced):
            print(f"    {note}")
    print(f"Wrote {os.path.relpath(OUT_CSV, ROOT)} "
          f"({len(kept)} poles > {cutoff:.0f} Ma x {len(header)} columns)")


if __name__ == "__main__":
    main()
